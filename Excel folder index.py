import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font



################# HERE YOU PUT THE ROUTE ######################
#ir could be a raw_input, but its more a one run code         #

route = 'C:/Your route here'

###############################################################

book = Workbook()
sheet = book.active
sheet.sheet_view.showGridLines = False

columnTemp = 2
rowTemp= 1

for subdir, dirs, files in os.walk(route):
    nCar=subdir.split('\\')
    folder = nCar
    columnTemp = len (folder)
    
    sheet.cell(row=rowTemp, column=columnTemp).font = Font(bold=True)
    sheet.cell(row=rowTemp, column=columnTemp).value = folder[-1]
    
    columnTemp = 1
    rowTemp += 1
    
    for i in range (len (folder)):
        columnTemp += 1
        
    for file in files:
        try:
            name,extension=file.split(".")
            
        except:
            fileTemp = file.split(".")
            name = ""
            extension = fileTemp[-1]
            
            for lugar in range(len(fileTemp)-1):
                if (lugar >= 1):
                    name +="."
                    
                name+=fileTemp[lugar]
                
        if (name.lower()=='thumbs'):
            break
        
        sheet.cell(row=rowTemp, column=1).value = extension.upper()
        sheet.cell(row=rowTemp, column=1).font = Font(bold=True)
        sheet.cell(row=rowTemp, column=columnTemp).value = name

        rowTemp += 1
   
    columnTemp = 1
    
    
################# HERE YOU PUT THE NAME OF THE FILE ###########

book.save("NAME OF THE FILE.xlsx")

###############################################################
